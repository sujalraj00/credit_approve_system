from django.core.management.base import BaseCommand
import openpyxl
from api.models import Customer, Loan
from datetime import datetime
import os

class Command(BaseCommand):
    help = "Load customer and loan data from Excel files"

    def handle(self, *args, **kwargs):
        # check if files are present or not
        customer_file = 'customer_data.xlsx'
        loan_file = 'loan_data.xlsx'
        
        if not os.path.exists(customer_file):
            self.stdout.write(self.style.ERROR(f"File {customer_file} not found"))
            return
        
        if not os.path.exists(loan_file):
            self.stdout.write(self.style.ERROR(f"File {loan_file} not found"))
            return

        # load customer data excle sheet
        self.stdout.write("Loading customer data...")
        customer_wb = openpyxl.load_workbook(customer_file)
        sheet = customer_wb.active
        
        customer_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # skip completely empty row
            if not any(row):
                continue
                
            # check for min req fields (assuming customer_id, first_name, last_name, phone_number are required)
            if not row[0] or not row[1] or not row[2] or not row[3]:
                self.stdout.write(f"Skipping invalid customer row: {row}")
                continue

            try:
                customer, created = Customer.objects.update_or_create(
                    id=row[0],  # use customer_id from the file as the primary key
                    defaults={
                        "first_name": str(row[1]).strip() if row[1] else "",
                        "last_name": str(row[2]).strip() if row[2] else "",
                        "phone_number": str(row[3]).strip() if row[3] else "",
                        "monthly_salary": float(row[4]) if row[4] else 0.0,
                        "approved_limit": float(row[5]) if row[5] else 0.0,
                        "current_debt": float(row[6]) if row[6] else 0.0,
                        "age": 30  
                    }
                )
                customer_count += 1
                if created:
                    self.stdout.write(f"Created customer: {customer.first_name} {customer.last_name}")
                else:
                    self.stdout.write(f"Updated customer: {customer.first_name} {customer.last_name}")
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Error processing customer row {row}: {e}"))

        self.stdout.write(f"Processed {customer_count} customers")

        # load loan data excel file 
        self.stdout.write("Loading loan data...")
        loan_wb = openpyxl.load_workbook(loan_file)
        sheet = loan_wb.active

        loan_count = 0
        skipped_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # skip completely empty rows
            if not any(row):
                continue
                
            # check for min req fields
            if not row[0] or not row[1]: 
                self.stdout.write(f"Skipping invalid loan row: {row}")
                skipped_count += 1
                continue

            try:
                # trying to get the customer
                customer = Customer.objects.get(id=row[0])
                
               
                start_date = None
                end_date = None
                
                if row[7]:  # start_date
                    if isinstance(row[7], datetime):
                        start_date = row[7].date()
                    else:
                        try:
                            # trying differ date formats
                            date_str = str(row[7]).split()[0]
                            start_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                start_date = datetime.strptime(date_str, "%m/%d/%Y").date()
                            except ValueError:
                                self.stdout.write(f"Warning: Could not parse start_date: {row[7]}")
                                start_date = datetime.now().date()
                
                if row[8]:  # end_date
                    if isinstance(row[8], datetime):
                        end_date = row[8].date()
                    else:
                        try:
                            # trying  different date formats
                            date_str = str(row[8]).split()[0]
                            end_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        except ValueError:
                            try:
                                end_date = datetime.strptime(date_str, "%m/%d/%Y").date()
                            except ValueError:
                                self.stdout.write(f"Warning: Could not parse end_date: {row[8]}")
                                end_date = start_date if start_date else datetime.now().date()

                # create or update loan - using loan_id as unique identifier
                loan, created = Loan.objects.update_or_create(
                    id=row[1],  # Use loan_id from the file as the primary key
                    defaults={
                        "customer": customer,
                        "loan_amount": float(row[2]) if row[2] else 0.0,
                        "tenure": int(row[3]) if row[3] else 1,
                        "interest_rate": float(row[4]) if row[4] else 0.0,
                        "monthly_repayment": float(row[5]) if row[5] else 0.0,
                        "emis_paid_on_time": int(row[6]) if row[6] else 0,
                        "start_date": start_date,
                        "end_date": end_date,
                    }
                )
                loan_count += 1
                
                if created:
                    self.stdout.write(f"Created loan {loan.id} for customer {customer.first_name} {customer.last_name}")
                else:
                    self.stdout.write(f"Updated loan {loan.id} for customer {customer.first_name} {customer.last_name}")
                    
            except Customer.DoesNotExist:
                self.stdout.write(f"Warning: Customer with ID {row[0]} not found, skipping loan {row[1]}")
                skipped_count += 1
                continue
            except Exception as e:
                self.stdout.write(self.style.ERROR(f"Error processing loan row {row}: {e}"))
                skipped_count += 1

        self.stdout.write(f"Processed {loan_count} loans")
        self.stdout.write(f"Skipped {skipped_count} loans")
        self.stdout.write(self.style.SUCCESS("Excel data loaded successfully!"))
        
        
        total_customers = Customer.objects.count()
        total_loans = Loan.objects.count()
        self.stdout.write(f"Total customers in database: {total_customers}")
        self.stdout.write(f"Total loans in database: {total_loans}")